import streamlit as st
import pandas as pd
from datetime import date
from reloj_circular import generar_reloj
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# =========================================================
# Configuración general
# =========================================================
st.set_page_config(page_title="Reloj de Tiempos Muertos", layout="wide")
st.title("📊 Reloj Circular de Tiempos Muertos")

# =========================================================
# CONFIGURACIÓN GOOGLE SHEETS (XLSX export)
# =========================================================
SHEET_EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1GSoaEg-ZUn5jB_VvLXcCkZjUnLR24ynIBPH3BcpCXXM/export?format=xlsx"
)

# =========================================================
# Carga de datos
# =========================================================
@st.cache_data(show_spinner=False)
def cargar_excel_desde_sheet(url: str) -> pd.DataFrame:
    df = pd.read_excel(url, sheet_name=0, engine="openpyxl", header=1)
    return df

def normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    rename_map = {}
    for c in df.columns:
        cl = (
            str(c)
            .strip()
            .lower()
            .replace("í", "i").replace("á", "a").replace("é", "e")
            .replace("ó", "o").replace("ú", "u")
        )
        if cl in ["fecha", "fecha y hora", "fecha/hora", "timestamp", "date", "datetime"]:
            rename_map[c] = "Fecha"
        if cl in ["id equipo", "id_equipo", "id maquina", "id máquina", "equipo", "machineid", "idequipo"]:
            rename_map[c] = "Id Equipo"
    df = df.rename(columns=rename_map)
    if "Fecha" not in df.columns or "Id Equipo" not in df.columns:
        raise ValueError("No se encuentran las columnas requeridas: 'Fecha' y 'Id Equipo'.")

    # --- Parseo robusto de fechas (soporta ISO, day-first y serial de Excel) ---
    col_fecha = df["Fecha"]

    if pd.api.types.is_datetime64_any_dtype(col_fecha):
        pass
    else:
        f0 = pd.to_datetime(col_fecha, errors="coerce", infer_datetime_format=True)
        mask_nat = f0.isna()
        if mask_nat.any():
            f1 = pd.to_datetime(
                col_fecha[mask_nat], errors="coerce", dayfirst=True, infer_datetime_format=True
            )
            f0.loc[mask_nat] = f1
        mask_nat = f0.isna()
        if mask_nat.any():
            try:
                numeric = pd.to_numeric(col_fecha[mask_nat], errors="coerce")
                have_num = numeric.notna()
                if have_num.any():
                    f2 = pd.to_datetime(
                        numeric[have_num],
                        unit="d",
                        origin="1899-12-30",
                        errors="coerce",
                    )
                    idx_nat = col_fecha[mask_nat].index
                    idx_fill = idx_nat[have_num]
                    f0.loc[idx_fill] = f2
            except Exception:
                pass
        df["Fecha"] = f0

    df["Id Equipo"] = df["Id Equipo"].astype(str).str.strip()
    return df

with st.spinner("Cargando datos de Google Sheets..."):
    raw = cargar_excel_desde_sheet(SHEET_EXPORT_URL)

try:
    df = normalizar_columnas(raw)
except Exception as e:
    st.error(f"Error al interpretar columnas: {e}")
    st.stop()

# =========================================================
# Interfaz
# =========================================================
maquinas = sorted(df["Id Equipo"].dropna().unique())
if not maquinas:
    st.error("No se encontraron máquinas en el archivo.")
    st.stop()

col_top1, col_top2, col_top3 = st.columns([1, 1, 1])
maquina_sel = col_top1.selectbox("Máquina", maquinas, index=0)

fechas_disponibles = sorted(pd.Series(df["Fecha"].dt.date.dropna().unique()).tolist())
if not fechas_disponibles:
    st.error("No se encontraron fechas en el archivo.")
    st.stop()

modo_multiple = col_top2.toggle("Seleccionar múltiples fechas", value=False, help="Activá para elegir más de una fecha")

if not modo_multiple:
    fecha_defecto = fechas_disponibles[0]
    fecha_sel = col_top3.selectbox(
        "Fecha",
        fechas_disponibles,
        index=fechas_disponibles.index(fecha_defecto) if fechas_disponibles else 0,
    )
    fechas_seleccionadas = [fecha_sel]
else:
    preselect = fechas_disponibles[-5:] if len(fechas_disponibles) >= 5 else fechas_disponibles
    fechas_seleccionadas = col_top3.multiselect(
        "Fechas (podés elegir varias)",
        options=fechas_disponibles,
        default=preselect,
        help="Elegí una o más fechas. Se generará un gráfico por cada día seleccionado.",
    )
    if not fechas_seleccionadas:
        st.info("Seleccioná al menos una fecha para continuar.")
        st.stop()

umbral_min = st.number_input(
    "Umbral de pausa no planificada (min)",
    min_value=1, max_value=30, value=3, step=1,
    help="Solo se consideran tiempos muertos con duración mayor o igual a este umbral."
)

def fmt_hms_from_timedelta(td: pd.Timedelta) -> str:
    total = int(td.total_seconds())
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def render_dia(fecha_dia):
    st.subheader(f"📅 Día {fecha_dia}")
    with st.spinner("Procesando..."):
        fig, indicadores, lista_gaps = generar_reloj(
            df, maquina_sel, fecha_dia, umbral_minutos=umbral_min
        )

    st.pyplot(fig, use_container_width=True)

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total disponible (min)", f"{indicadores['total_disponible']:.2f}")
    c2.metric("Inutilizado (pausas, min)", f"{indicadores['inutilizado_programado']:.2f}")
    c3.metric("Neto (min)", f"{indicadores['neto']:.2f}")
    c4.metric("Perdido no programado (min)", f"{indicadores['perdido_no_programado']:.2f}")
    c5.metric("% Perdido", f"{indicadores['porcentaje_perdido']:.2f}")

    st.markdown(
        f"**Tiempos muertos detectados (≥ {umbral_min} min):** "
        f"**{len(lista_gaps)} intervalos**, total **{indicadores['perdido_no_programado']:.2f} min**."
    )

    if not lista_gaps:
        st.info("No se detectaron tiempos muertos para este día.")
        return None

    df_gaps = pd.DataFrame(lista_gaps)
    td = pd.to_timedelta(df_gaps["Duracion_min"], unit="m")
    df_show = df_gaps.copy()
    df_show["Duracion"] = td.apply(fmt_hms_from_timedelta)
    df_show = df_show[["Inicio", "Fin", "Duracion"]]
    st.dataframe(df_show, use_container_width=True)

    csv_df = df_show.copy()
    csv_bytes = csv_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "📥 Descargar detalle (CSV)",
        data=csv_bytes,
        file_name=f"tiempos_muertos_{maquina_sel}_{fecha_dia}.csv",
        mime="text/csv",
        use_container_width=True,
    )

    from io import BytesIO
    from openpyxl.utils import get_column_letter
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_xlsx = df_gaps.copy()
        df_xlsx["Duracion"] = pd.to_timedelta(df_xlsx["Duracion_min"], unit="m").dt.total_seconds() / 86400.0
        df_xlsx = df_xlsx[["Inicio", "Fin", "Duracion"]]
        df_xlsx.to_excel(writer, index=False, sheet_name="TiemposMuertos")
        ws = writer.book["TiemposMuertos"]
        dur_col_letter = get_column_letter(3)
        for row in range(2, ws.max_row + 1):
            ws[f"{dur_col_letter}{row}"].number_format = "[h]:mm:ss"
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12

    st.download_button(
        "📥 Descargar detalle (Excel)",
        data=output.getvalue(),
        file_name=f"tiempos_muertos_{maquina_sel}_{fecha_dia}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    return {
        "Fecha": fecha_dia,
        "Total_disponible_min": indicadores["total_disponible"],
        "Inutilizado_prog_min": indicadores["inutilizado_programado"],
        "Neto_min": indicadores["neto"],
        "Perdido_no_prog_min": indicadores["perdido_no_programado"],
        "%_Perdido": indicadores["porcentaje_perdido"],
    }

if st.button("Generar gráfico(s)", type="primary", use_container_width=True):
    res_resumen = []
    for f in fechas_seleccionadas:
        with st.container():
            resumen = render_dia(f)
            st.divider()
            if resumen:
                res_resumen.append(resumen)

    if len(res_resumen) > 1:
        st.subheader("📈 Resumen de días seleccionados")
        df_res = pd.DataFrame(res_resumen)
        df_res_display = df_res.copy()
        for col in ["Total_disponible_min", "Inutilizado_prog_min", "Neto_min", "Perdido_no_prog_min", "%_Perdido"]:
            df_res_display[col] = df_res_display[col].map(lambda x: f"{x:.2f}")
        st.dataframe(df_res_display, use_container_width=True)

        # ========= Gráfico de línea histórico (% Perdido) =========
        st.markdown("#### 📉 Histórico de % Perdido")
        df_plot = df_res.copy()
        df_plot["Fecha_dt"] = pd.to_datetime(df_plot["Fecha"])
        df_plot = df_plot.groupby("Fecha_dt", as_index=False)["%_Perdido"].mean()
        df_plot = df_plot.sort_values("Fecha_dt")

        fig, ax = plt.subplots(figsize=(8, 3))
        ax.plot(df_plot["Fecha_dt"], df_plot["%_Perdido"], marker="o", linewidth=2)
        ax.set_xlabel("Fecha")
        ax.set_ylabel("% Perdido")

        # Escala dinámica del eje Y (doble del valor máximo)
        max_y = df_plot["%_Perdido"].max() if not df_plot.empty else 0
        ax.set_ylim(bottom=0, top=max_y * 2)

        # ✅ Mostrar SOLO las fechas con datos (sin rellenar días faltantes)
        ax.set_xticks(df_plot["Fecha_dt"])
        ax.set_xticklabels(df_plot["Fecha_dt"].dt.strftime("%Y-%m-%d"), rotation=45, ha="right")

        ax.grid(True, alpha=0.3)
        st.pyplot(fig, use_container_width=True)
        # ========= FIN gráfico histórico =========

        csv_bytes = df_res.to_csv(index=False).encode("utf-8")
        st.download_button(
            "📥 Descargar resumen (CSV)",
            data=csv_bytes,
            file_name=f"resumen_{maquina_sel}.csv",
            mime="text/csv",
            use_container_width=True,
        )

        from io import BytesIO
        output_sum = BytesIO()
        with pd.ExcelWriter(output_sum, engine="openpyxl") as writer:
            df_res.to_excel(writer, index=False, sheet_name="Resumen")
        st.download_button(
            "📥 Descargar resumen (Excel)",
            data=output_sum.getvalue(),
            file_name=f"resumen_{maquina_sel}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
